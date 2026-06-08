import csv
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def env_path(name, default):
    value = os.environ.get(name, "")
    return value if value else default

base_dir = env_path("UBL3_ROUTE_C_BASE_DIR", r"D:\RNA\2026063Molecular Neurodegeneration")
fig4_dir = env_path(
    "UBL3_FIG4_SOURCE_DIR",
    os.path.join(base_dir, "Figure4", "results", "RouteC_20260606_v26_2x2_axis_title_spacing")
)
out_dir = env_path(
    "UBL3_SUPPTABLE_S2_OUT_DIR",
    os.path.join(base_dir, "Supplementary_Table_S2", "results", "RouteC_20260606_robustness_sensitivity_aligned_to_Fig4_v26")
)
os.makedirs(out_dir, exist_ok=True)

def read_csv_dict(name):
    path = os.path.join(fig4_dir, name)
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    with open(path, newline='', encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f))
    return rows

def rows_from_dicts(rows):
    if not rows:
        return [[]]
    headers = list(rows[0].keys())
    return [headers] + [[r.get(h, '') for h in headers] for r in rows]

def write_sheet(wb, name, rows):
    ws = wb.create_sheet(name)
    for row in rows:
        ws.append(row)
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.freeze_panes = 'A2'
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            val = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(val), 80))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 55))
    return ws

panelA = read_csv_dict("Figure4_panelA_multiple_testing_source.csv")
panelB = read_csv_dict("Figure4_panelB_detection_count_source.csv")
panelC_sum = read_csv_dict("Figure4_panelC_leave_one_out_summary.csv")
panelC_iter = read_csv_dict("Figure4_panelC_leave_one_out_source.csv")
panelD = read_csv_dict("Figure4_panelD_subcluster_source.csv")

readme = [
    ["Item", "Description"],
    ["Workbook title", "Supplementary Table S2. Robustness and sensitivity analyses for the PSP V1 cortical-neuron UBL3 candidate signal."],
    ["Companion figure", "Figure 4, Route C final v26."],
    ["Purpose", "Source tables for robustness and sensitivity checks of PSP V1 excitatory- and inhibitory-neuron UBL3 detection-breadth findings."],
    ["Statistical unit", "Donor is the statistical unit for Wilcoxon/Hodges-Lehmann analyses; detection-count models use donor-level binomial counts with quasibinomial dispersion."],
    ["Panel A", "Multiple-testing sensitivity for the two PSP V1 cortical-neuron findings across increasingly broad correction families."],
    ["Panel B", "Detection-count quasibinomial sensitivity using UBL3-positive nuclei counts and total nuclei per donor; reports odds ratios, 95% confidence intervals, raw P values and q values."],
    ["Panel C", "Leave-one-donor-out donor-level robustness for the two PSP V1 cortical-neuron findings."],
    ["Panel D", "Exploratory source-label neuronal subcluster localization within PSP V1 excitatory and inhibitory neurons; this does not replace the six-class primary framework."],
    ["Multiplicity note", "Within-unit q values refer to BH correction across the six major cell classes in the relevant disease-region unit unless otherwise stated; broader correction families are shown as sensitivity checks."],
    ["Source files", fig4_dir],
    ["Created", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
]

workbook_map = [
    ["sheet_name", "figure_panel", "source_file", "description"],
    ["README", "", "", "Workbook description and definitions."],
    ["Workbook_map", "", "", "Sheet map."],
    ["PanelA_multiple_testing", "Figure 4A", "Figure4_panelA_multiple_testing_source.csv", "Multiple-testing sensitivity results."],
    ["PanelB_detection_count", "Figure 4B", "Figure4_panelB_detection_count_source.csv", "Detection-count quasibinomial model results."],
    ["PanelC_LOO_summary", "Figure 4C", "Figure4_panelC_leave_one_out_summary.csv", "Leave-one-donor-out summary."],
    ["PanelC_LOO_iterations", "Figure 4C", "Figure4_panelC_leave_one_out_source.csv", "Leave-one-donor-out iteration-level results."],
    ["PanelD_subclusters", "Figure 4D", "Figure4_panelD_subcluster_source.csv", "Exploratory neuronal subcluster localization results."],
    ["Source_manifest", "", "", "Input file paths and sizes."],
]

source_files = [
    "Figure4_panelA_multiple_testing_source.csv",
    "Figure4_panelB_detection_count_source.csv",
    "Figure4_panelC_leave_one_out_summary.csv",
    "Figure4_panelC_leave_one_out_source.csv",
    "Figure4_panelD_subcluster_source.csv",
    "Figure4_RouteC_count_subclusters_legend_draft.txt",
    "Figure4_RouteC_count_subclusters_file_sizes.csv",
]
source_manifest = [["file_name", "path", "exists", "size_bytes"]]
for fn in source_files:
    p = os.path.join(fig4_dir, fn)
    source_manifest.append([fn, p, os.path.exists(p), os.path.getsize(p) if os.path.exists(p) else ""])

wb = Workbook()
wb.remove(wb.active)
write_sheet(wb, "README", readme)
write_sheet(wb, "Workbook_map", workbook_map)
write_sheet(wb, "PanelA_multiple_testing", rows_from_dicts(panelA))
write_sheet(wb, "PanelB_detection_count", rows_from_dicts(panelB))
write_sheet(wb, "PanelC_LOO_summary", rows_from_dicts(panelC_sum))
write_sheet(wb, "PanelC_LOO_iterations", rows_from_dicts(panelC_iter))
write_sheet(wb, "PanelD_subclusters", rows_from_dicts(panelD))
write_sheet(wb, "Source_manifest", source_manifest)

out_xlsx = os.path.join(out_dir, "Supplementary_Table_S2_robustness_sensitivity_RouteC_aligned_to_Figure4_v26.xlsx")
wb.save(out_xlsx)

# Validate with openpyxl readback and ensure no filled cells were introduced.
wb2 = load_workbook(out_xlsx, read_only=False, data_only=True)
for ws in wb2.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fill_type not in (None, 'none'):
                raise RuntimeError(f"Unexpected fill in {ws.title}!{cell.coordinate}")
print("WROTE", out_xlsx)
print("SHEETS", ", ".join(wb2.sheetnames))
print("SIZE_MB", round(os.path.getsize(out_xlsx)/1024/1024, 4))